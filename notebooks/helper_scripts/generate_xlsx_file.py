import json
import openpyxl
import os
import scipy
import librosa

data_path = os.path.dirname(os.path.abspath(__file__))
folders = [os.path.join(data_path, name) for name in os.listdir(data_path)
           if os.path.isdir(os.path.join(data_path, name))]
json_files = [os.path.join(data_path, folder, file) for folder in folders for file in os.listdir(folder)
              if file.endswith('.json')]
audio_files = [os.path.join(data_path, folder, file) for folder in folders for file in os.listdir(folder)
               if not file.endswith('.json')]

data = []
keys = ["id", "cnt", "type", "q", "time"]

for json_file in json_files:
    with open(json_file, 'r') as f:
        parsed_json = json.load(f)
        for recording in parsed_json["recordings"]:
            all_keys = list(recording.keys())
            for key in all_keys:
                if key not in keys:
                    del recording[key]
            data.append(recording)

wb = openpyxl.Workbook()
sheet = wb['Sheet']
n = 1
for key in keys:
    sheet.cell(row=1, column=n).value = key
    n += 1
r = 2
for record in data:
    c = 1
    for key in keys:
        sheet.cell(row=r, column=c).value = record[key]
        c += 1
    r += 1
wb.save("data.xlsx")
